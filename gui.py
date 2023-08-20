from openpyxl import Workbook

# 创建一个字典
data = {'Name': ['Alice', 'Bob', 'Charlie'],
        'Age': [25, 30, 35],
        'City': ['New York', 'London', 'Paris']}

# 创建一个Workbook对象
workbook = Workbook()

# 获取默认的Sheet
sheet = workbook.active

# 写入表头
header = list(data.keys())
for col_num, header_value in enumerate(header, 1):
    sheet.cell(row=1, column=col_num).value = header_value

# 写入数据
for row_num, values in enumerate(zip(*data.values()), 2):
    for col_num, value in enumerate(values, 1):
        sheet.cell(row=row_num, column=col_num).value = value

# 保存Excel文件
workbook.save('output.xlsx')